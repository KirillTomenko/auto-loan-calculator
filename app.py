"""
Flask приложение для кредитного калькулятора.
"""

import os
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from calculator import calculate_loan
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Создаём директории для экспортов
os.makedirs('exports', exist_ok=True)
os.makedirs('static/css', exist_ok=True)
os.makedirs('static/js', exist_ok=True)
os.makedirs('templates', exist_ok=True)


@app.route('/')
def index():
    """Главная страница."""
    return render_template('index.html')


@app.route('/api/calculate', methods=['POST'])
def api_calculate():
    """API endpoint для расчёта кредита."""
    try:
        data = request.json
        
        principal = float(data.get('principal', 0))
        rate = float(data.get('rate', 0))
        term_months = int(data.get('term_months', 0))
        start_date = data.get('start_date')
        early_payments = data.get('early_payments', {})
        
        # Конвертируем early_payments в правильный формат
        formatted_early_payments = {}
        if early_payments:
            for month_str, payment_data in early_payments.items():
                month = int(month_str)
                formatted_early_payments[month] = {
                    'amount': float(payment_data['amount']),
                    'mode': payment_data.get('mode', 'reduce_payment')
                }
        
        result = calculate_loan(
            principal=principal,
            rate=rate,
            term_months=term_months,
            start_date=start_date,
            early_payments=formatted_early_payments if formatted_early_payments else None
        )
        
        return jsonify(result)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/export', methods=['POST'])
def api_export():
    """API endpoint для экспорта в Excel."""
    try:
        data = request.json
        
        parameters = data.get('parameters', {})
        schedule = data.get('schedule', [])
        
        # Генерируем имя файла
        today = datetime.now()
        filename = f"автокредит_{today.strftime('%d%m%y')}.xlsx"
        filepath = os.path.join('exports', filename)
        
        export_to_excel(parameters, schedule, filepath)
        
        return jsonify({
            'filename': filename,
            'url': f'/downloads/{filename}'
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/downloads/<filename>')
def download_file(filename):
    """Скачивание Excel файла."""
    filepath = os.path.join('exports', filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': 'File not found'}), 404


def export_to_excel(parameters: dict, schedule: list, filepath: str):
    """
    Экспортирует расчёты в Excel-файл с тремя листами.
    
    Args:
        parameters: Словарь с параметрами кредита
        schedule: Список словарей с графиком платежей
        filepath: Путь для сохранения файла
    """
    wb = Workbook()
    
    # Удаляем дефолтный лист
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Лист 1: Исходные параметры
    ws_params = wb.create_sheet('Параметры кредита')
    ws_params.append(['Параметр', 'Значение'])
    
    params_data = [
        ['Стоимость автомобиля', parameters.get('principal', 0)],
        ['Первоначальный взнос', parameters.get('down_payment', 0)],
        ['Сумма кредита', parameters.get('principal', 0)],
        ['Процентная ставка (%)', parameters.get('rate', 0)],
        ['Срок кредита (месяцев)', parameters.get('term_months', 0)],
        ['Дата получения кредита', parameters.get('start_date', '')],
    ]
    
    for row in params_data:
        ws_params.append(row)
    
    # Стилизация листа параметров
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    for cell in ws_params[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row in ws_params.iter_rows(min_row=2, max_row=len(params_data) + 1):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if cell.column == 2:  # Колонка значений
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
    
    ws_params.column_dimensions['A'].width = 30
    ws_params.column_dimensions['B'].width = 20
    
    # Лист 2: График платежей
    ws_schedule = wb.create_sheet('График платежей')
    headers = [
        '№ месяца', 'Дата платежа', 'Ежемесячный платёж',
        'Досрочный платёж', 'Выплачено основной суммы',
        'Выплачено процентов', 'Оставшийся долг'
    ]
    ws_schedule.append(headers)
    
    # Стилизация заголовков
    for cell in ws_schedule[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Добавляем данные
    for payment in schedule:
        row = [
            payment.get('month', ''),
            payment.get('payment_date', ''),
            payment.get('monthly_payment', 0),
            payment.get('early_payment', 0),
            payment.get('principal_paid', 0),
            payment.get('interest_paid', 0),
            payment.get('remaining_balance', 0)
        ]
        ws_schedule.append(row)
    
    # Форматирование чисел
    for row in ws_schedule.iter_rows(min_row=2, max_row=len(schedule) + 1):
        for idx, cell in enumerate(row):
            if idx >= 2:  # Начиная с колонки "Ежемесячный платёж"
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right' if idx >= 2 else 'center', vertical='center')
    
    # Итоговая строка
    total_row = [
        'ИТОГО',
        '',
        sum(p.get('monthly_payment', 0) for p in schedule),
        sum(p.get('early_payment', 0) for p in schedule),
        sum(p.get('principal_paid', 0) for p in schedule),
        sum(p.get('interest_paid', 0) for p in schedule),
        0
    ]
    ws_schedule.append(total_row)
    
    # Стилизация итоговой строки
    total_fill = PatternFill(start_color='D1D5DB', end_color='D1D5DB', fill_type='solid')
    total_font = Font(bold=True, size=11)
    for cell in ws_schedule[ws_schedule.max_row]:
        cell.fill = total_fill
        cell.font = total_font
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'
    
    # Настройка ширины колонок
    column_widths = [12, 15, 18, 18, 20, 18, 18]
    for idx, width in enumerate(column_widths, start=1):
        ws_schedule.column_dimensions[get_column_letter(idx)].width = width
    
    # Лист 3: Итоговая сводка
    ws_summary = wb.create_sheet('Итоговая сводка')
    
    # Подсчитываем итоги
    total_interest = sum(p.get('interest_paid', 0) for p in schedule)
    total_principal = sum(p.get('principal_paid', 0) for p in schedule)
    total_early = sum(p.get('early_payment', 0) for p in schedule)
    monthly_payment = schedule[0].get('monthly_payment', 0) if schedule else 0
    
    summary_data = [
        ['Показатель', 'Значение'],
        ['Ежемесячный платёж', monthly_payment],
        ['Общая переплата по процентам', total_interest],
        ['Общая сумма досрочных платежей', total_early],
        ['Итого выплачено основной суммы', total_principal],
        ['Итого к выплате', total_principal + total_interest + total_early],
        ['', ''],
        ['Примерный необходимый ежемесячный доход', monthly_payment * 2.5],
    ]
    
    for row in summary_data:
        ws_summary.append(row)
    
    # Стилизация заголовка
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Форматирование значений
    for row in ws_summary.iter_rows(min_row=2, max_row=len(summary_data)):
        for cell in row:
            if cell.column == 2 and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            cell.alignment = Alignment(
                horizontal='left' if cell.column == 1 else 'right',
                vertical='center'
            )
    
    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 25
    
    # Сохраняем файл
    wb.save(filepath)


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
